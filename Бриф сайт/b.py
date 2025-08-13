# pip install openpyxl matplotlib
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from io import BytesIO

# ------------------------------------------------------------------------------
# Настройки темы (дружелюбная светлая палитра, мягкие оттенки, хороший контраст)
# ------------------------------------------------------------------------------
theme = {
    "bg": "F7F9FC",            # общий светлый фон
    "card": "FFFFFF",          # карточки блоков
    "primary": "3B82F6",       # акцент (синий)
    "primary_dark": "1D4ED8",
    "text": "1F2937",          # основной текст (тёмно-серый)
    "muted": "6B7280",         # вторичный текст
    "line": "E5E7EB",          # линии/границы
    "success": "10B981",       # зелёный (успех)
    "warning": "F59E0B",       # жёлто-оранжевый (внимание)
    "info": "0EA5E9",          # голубой (инфо)
    "badge_bg": "EFF6FF",      # фон бейджей
    "badge_text": "1E3A8A"     # текст бейджей
}

# ------------------------------------------------------------------------------
# Создание книги и листа
# ------------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Affort_Brief"

# Общий фон: имитируем фоном заливкой крупных областей
for r in range(1, 200):
    for c in range(1, 12):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=theme["bg"])

# ------------------------------------------------------------------------------
# Удобочитаемая типографика (стили)
# ------------------------------------------------------------------------------
title_style = NamedStyle(name="title_style")
title_style.font = Font(size=18, bold=True, color=theme["primary_dark"])
title_style.alignment = Alignment(horizontal="center", vertical="center")

section_title_style = NamedStyle(name="section_title_style")
section_title_style.font = Font(size=13, bold=True, color=theme["text"])
section_title_style.alignment = Alignment(vertical="center")
section_title_style.fill = PatternFill("solid", fgColor=theme["card"])

label_style = NamedStyle(name="label_style")
label_style.font = Font(size=11, color=theme["muted"])
label_style.alignment = Alignment(vertical="center", wrap_text=True)
label_style.fill = PatternFill("solid", fgColor=theme["card"])

input_style = NamedStyle(name="input_style")
input_style.font = Font(size=11, color=theme["text"])
input_style.alignment = Alignment(vertical="center", wrap_text=True)
input_style.fill = PatternFill("solid", fgColor="FFFFFF")

badge_style = NamedStyle(name="badge_style")
badge_style.font = Font(size=10, bold=True, color=theme["badge_text"])
badge_style.alignment = Alignment(horizontal="center", vertical="center")
badge_style.fill = PatternFill("solid", fgColor=theme["badge_bg"])

if "title_style" not in wb.named_styles:
    wb.add_named_style(title_style)
if "section_title_style" not in wb.named_styles:
    wb.add_named_style(section_title_style)
if "label_style" not in wb.named_styles:
    wb.add_named_style(label_style)
if "input_style" not in wb.named_styles:
    wb.add_named_style(input_style)
if "badge_style" not in wb.named_styles:
    wb.add_named_style(badge_style)

# Универсальные границы (мягкие)
thin_border = Border(
    left=Side(style='thin', color=theme["line"]),
    right=Side(style='thin', color=theme["line"]),
    top=Side(style='thin', color=theme["line"]),
    bottom=Side(style='thin', color=theme["line"])
)

card_border = Border(
    left=Side(style='medium', color=theme["bg"]),
    right=Side(style='medium', color=theme["bg"]),
    top=Side(style='medium', color=theme["bg"]),
    bottom=Side(style='medium', color=theme["bg"])
)

# ------------------------------------------------------------------------------
# Адаптивная сетка (ширины колонок, «воздух»)
# ------------------------------------------------------------------------------
col_widths = {
    1: 4,   # отступ
    2: 34,  # метки
    3: 4,   # разделитель
    4: 18,  # поле
    5: 18,  # поле
    6: 18,  # поле
    7: 18,  # поле
    8: 4,   # отступ/иконки
    9: 26,  # правый сайдбар/визуал
    10: 4,  # отступ
    11: 4   # отступ
}
for idx, w in col_widths.items():
    ws.column_dimensions[get_column_letter(idx)].width = w

# ------------------------------------------------------------------------------
# Упрощённая вводная (коротко и по делу)
# ------------------------------------------------------------------------------
row = 2
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
cell = ws.cell(row=row, column=2)
cell.value = "Бриф для международной бизнес-миссии — Affort"
cell.style = "title_style"
cell.fill = PatternFill("solid", fgColor=theme["bg"])
ws.row_dimensions[row].height = 34
row += 2

ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=7)
intro = ws.cell(row=row, column=2)
intro.value = (
    "Расскажите кратко о целях и ограничениях — мы подберём компании, маршруты и формат визитов, "
    "чтобы ваши руководители получили максимум инсайтов и партнёрств за минимум времени."
)
intro.font = Font(size=11, color=theme["muted"])
intro.alignment = Alignment(wrap_text=True, vertical="center")
intro.fill = PatternFill("solid", fgColor=theme["bg"])
ws.row_dimensions[row].height = 28
ws.row_dimensions[row+1].height = 22
row += 3

# Наглядные бейджи-фокусы (кратко, дружелюбно)
badges = ["Аналитика", "Визиты", "Обучение", "VIP/Логистика"]
for i, b in enumerate(badges):
    c = ws.cell(row=row, column=2+i)
    c.value = b
    c.style = "badge_style"
    c.border = thin_border
ws.row_dimensions[row].height = 24
row += 2

# ------------------------------------------------------------------------------
# Карточки секций (укороченные, дружелюбные метки, оптимизированная структура)
# ------------------------------------------------------------------------------
sections = [
    {
        "title": "1. О компании и контакт",
        "fields": [
            "Название компании",
            "Сфера деятельности",
            "Контакт (ФИО, роль, email, телефон)"
        ]
    },
    {
        "title": "2. Цели и ожидаемые результаты",
        "fields": [
            "Главная цель миссии (1-2 предложения)",
            "Ключевые задачи (3-5 пунктов)",
            "Ожидаемые KPI/результаты и сроки"
        ]
    },
    {
        "title": "3. Отрасли и география",
        "fields": [
            "Отрасли интереса (B2B/B2C, конкретика)",
            "Страны/города приоритетно",
            "Предварительное обоснование выбора"
        ]
    },
    {
        "title": "4. Делегация и бюджет",
        "fields": [
            "Состав и роли (кол-во C-level/экспертов)",
            "Предпочтительные даты и длительность",
            "Бюджетный коридор и формат оплаты"
        ]
    },
    {
        "title": "5. Формат и сервис",
        "fields": [
            "Формат визитов (демо/воркшоп/переговоры)",
            "Языки и перевод",
            "Размещение, безопасность, VIP/протокол"
        ]
    }
]

def draw_card(start_row, title, fields):
    # рамка карточки
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=7)
    header = ws.cell(row=start_row, column=2)
    header.value = title
    header.style = "section_title_style"
    header.fill = PatternFill("solid", fgColor=theme["card"])
    header.border = thin_border
    ws.row_dimensions[start_row].height = 24

    r = start_row + 1
    for f in fields:
        # метка
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        lab = ws.cell(row=r, column=2)
        lab.value = f
        lab.style = "label_style"
        lab.border = thin_border

        # поля ввода (две строки для «воздуха»)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        inp = ws.cell(row=r, column=4)
        inp.value = ""
        inp.style = "input_style"
        inp.border = thin_border
        ws.row_dimensions[r].height = 22
        r += 1

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        inp2 = ws.cell(row=r, column=4)
        inp2.value = ""
        inp2.style = "input_style"
        inp2.border = thin_border
        ws.row_dimensions[r].height = 22
        r += 1

    # нижний отступ
    r += 1
    return r

for sec in sections:
    row = draw_card(row, sec["title"], sec["fields"])

# ------------------------------------------------------------------------------
# Компактный дружелюбный блок «Как заполнять» (с юмором для топ-менеджмента)
# ------------------------------------------------------------------------------
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
tips_title = ws.cell(row=row, column=2)
tips_title.value = "Подсказки по заполнению (коротко):"
tips_title.font = Font(size=12, bold=True, color=theme["text"])
tips_title.fill = PatternFill("solid", fgColor=theme["bg"])
ws.row_dimensions[row].height = 22
row += 1

ws.merge_cells(start_row=row, start_column=2, end_row=row+2, end_column=7)
tips = ws.cell(row=row, column=2)
tips.value = (
    "— Пишите по существу: цель, 3-5 задач, 2-3 KPI. "
    "— Чем точнее даты и бюджет, тем быстрее маршрут. "
    "— Не бойтесь амбиций: под амбиции и составим миссию. "
    "— Вопросы «с подвохом» приветствуются — скучно не будет."
)
tips.font = Font(size=10, color=theme["muted"])
tips.alignment = Alignment(wrap_text=True, vertical="top")
tips.fill = PatternFill("solid", fgColor=theme["bg"])
ws.row_dimensions[row].height = 36
row += 3

# ------------------------------------------------------------------------------
# Лёгкая визуализация (минимализм, светлая тема)
# ------------------------------------------------------------------------------
plt.rcParams.update({"figure.facecolor": "#FFFFFF"})
fig, ax = plt.subplots(figsize=(6.2, 3.4))
categories = ["Аналитика", "Визиты", "Обучение", "VIP/Логистика"]
values = [30, 30, 20, 20]
bar_colors = ["#93C5FD", "#86EFAC", "#FDE68A", "#C7D2FE"]

bars = ax.bar(categories, values, color=bar_colors, edgecolor="#E5E7EB", linewidth=1.2)
ax.set_ylim(0, 40)
ax.set_ylabel("Доля фокуса (%)", color="#374151")
ax.set_title("Фокус миссии Affort (пример распределения)", color="#111827", fontsize=12, pad=10)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)
ax.spines["left"].set_color("#E5E7EB")
ax.spines["bottom"].set_color("#E5E7EB")
ax.tick_params(colors="#374151")

for bar, v in zip(bars, values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1, f"{v}%", ha="center", va="bottom", color="#111827", fontsize=10)

plt.tight_layout()
buf = BytesIO()
plt.savefig(buf, format="png", dpi=150)
buf.seek(0)
img = Image(buf)
img.width = 420
img.height = 230
ws.add_image(img, "I3")
plt.close()

# ------------------------------------------------------------------------------
# Короткий дружелюбный call-to-action
# ------------------------------------------------------------------------------
row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=7)
cta = ws.cell(row=row, column=2)
cta.value = (
    "Заполните бриф — и мы вернём вам маршрут, компании и формат встреч быстрее, "
    "чем вы успеете сказать «ROI». Пожалуйста, максимально конкретно: "
    "уточните цели, сроки и ограничения. Остальное — берём на себя."
)
cta.font = Font(size=11, color=theme["primary_dark"])
cta.alignment = Alignment(wrap_text=True, vertical="center")
cta.fill = PatternFill("solid", fgColor=theme["badge_bg"])
ws.row_dimensions[row].height = 38
ws.row_dimensions[row+1].height = 22
row += 3

# ------------------------------------------------------------------------------
# Пара косметических штрихов: выравнивания, высоты, тонкие границы для заполненных
# ------------------------------------------------------------------------------
max_row = row + 2
for r in range(1, max_row):
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 20, 20)
    for c in range(1, 12):
        cell = ws.cell(r, c)
        # тонкие рамки только на «карточках» и полях
        if cell.fill.fgColor.rgb in [theme["card"], "FFFFFF", theme["badge_bg"]]:
            cell.border = thin_border

# ------------------------------------------------------------------------------
# Сохранение
# ------------------------------------------------------------------------------
filename = "Affort_Brief_Friendly_Adaptive.xlsx"
wb.save(filename)
print(f"✅ Готово: {filename}")
